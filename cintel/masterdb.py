"""
Lesen und versioniertes Schreiben der Competitive Intel Master DB.

Kernprinzip: die Eingangsdatei wird NIE veraendert. Jeder Merge schreibt eine
neue Datei mit erhoehter Versionsnummer nach data/outputs/.
"""

from __future__ import annotations

import datetime as dt
import logging
import re
import shutil
from collections.abc import Iterator
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

from .schema import (
    COMPANY_LEVEL_FIELDS,
    ROW_KIND_COMPANY,
    Record,
    SchemaError,
    resolve_columns,
)

log = logging.getLogger(__name__)

SHEET_MASTER = "Competitors_All-Master"
SHEET_LEGEND = "Legend_Categories"

VERSION_RE = re.compile(r"_v(\d+)\.(\d+)\.xlsx$", re.IGNORECASE)


@dataclass
class CompanyBlock:
    """
    Alle Zeilen einer Firma: genau eine Company-Zeile plus N Produktzeilen.

    In v2.2 gibt es Firmen ohne Company-Zeile (nur Produktzeilen) - deshalb ist
    `company_row` optional.
    """

    company_id: Any
    company: str
    company_row: Record | None = None
    product_rows: list[Record] = field(default_factory=list)
    row_indices: list[int] = field(default_factory=list)

    @property
    def all_rows(self) -> list[Record]:
        rows: list[Record] = []
        if self.company_row is not None:
            rows.append(self.company_row)
        rows.extend(self.product_rows)
        return rows

    @property
    def urls(self) -> list[str]:
        out: list[str] = []
        for record in self.all_rows:
            value = record.values.get("url")
            if value and str(value).lower().startswith("http"):
                out.append(str(value).strip())
        return out

    @property
    def has_products(self) -> bool:
        return any(
            str(r.values.get("product_name") or "").strip()
            for r in self.product_rows
        )

    def missing_fields(self, fields: list[str]) -> list[str]:
        """
        Welche der genannten Felder fehlen - ebenenbewusst geprueft.

        Das ist wichtig, weil `Product_name` auf der Company-Zeile per Design
        leer ist. Eine naive Pruefung wuerde jede Firma als luecken behaftet
        melden.

          - Company-Level-Felder  -> auf der Company-Zeile pruefen
          - "product_name"        -> fehlt, wenn es gar keine Produktzeile gibt
          - sonstige Produktfelder-> fehlen, wenn KEINE Produktzeile sie fuellt
        """
        company_source = self.company_row or (
            self.product_rows[0] if self.product_rows else None
        )
        missing: list[str] = []

        for name in fields:
            if name == "product_name":
                if not self.has_products:
                    missing.append(name)
                continue

            if name in COMPANY_LEVEL_FIELDS:
                value = company_source.values.get(name) if company_source else None
                if not str(value or "").strip():
                    missing.append(name)
                continue

            # Produktebene: erfuellt, sobald mindestens eine Zeile den Wert hat.
            filled = any(
                str(r.values.get(name) or "").strip()
                for r in self.all_rows
            )
            if not filled:
                missing.append(name)

        return missing


class MasterDB:
    """Geladene Master-DB mit aufgeloester Spaltenzuordnung."""

    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)
        if not self.path.exists():
            raise FileNotFoundError(f"Master-DB nicht gefunden: {self.path}")

        workbook = openpyxl.load_workbook(self.path, data_only=True, read_only=True)
        if SHEET_MASTER not in workbook.sheetnames:
            raise SchemaError(
                f"Sheet '{SHEET_MASTER}' fehlt. Vorhanden: {workbook.sheetnames}"
            )
        sheet = workbook[SHEET_MASTER]

        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()

        if not rows:
            raise SchemaError("Master-DB ist leer.")

        self.headers: list[Any] = list(rows[0])
        self.column_map: dict[str, int] = resolve_columns(self.headers)
        self.records: list[Record] = []

        for raw in rows[1:]:
            if not any(v not in (None, "") for v in raw):
                continue
            values = {
                fieldname: (raw[pos] if pos < len(raw) else None)
                for fieldname, pos in self.column_map.items()
            }
            self.records.append(Record(values=values))

        log.info("Master-DB geladen: %d Zeilen, %d Spalten aus %s",
                 len(self.records), len(self.column_map), self.path.name)

    # -- Zugriff -----------------------------------------------------------

    def __len__(self) -> int:
        return len(self.records)

    def blocks(self) -> list[CompanyBlock]:
        """Gruppiert die Zeilen zu Firmenbloecken (Company-Zeile + Produkte)."""
        grouped: dict[str, CompanyBlock] = {}
        for index, record in enumerate(self.records):
            key = _block_key(record)
            if key not in grouped:
                grouped[key] = CompanyBlock(
                    company_id=record.values.get("company_id"),
                    company=str(record.values.get("company") or "").strip(),
                )
            block = grouped[key]
            block.row_indices.append(index)
            if record.values.get("row_kind") == ROW_KIND_COMPANY:
                if block.company_row is None:
                    block.company_row = record
                else:
                    block.product_rows.append(record)
            else:
                block.product_rows.append(record)
        return list(grouped.values())

    def company_names(self) -> set[str]:
        return {
            str(r.values.get("company") or "").strip()
            for r in self.records
            if r.values.get("company")
        }

    def next_company_id(self) -> int:
        """Naechste freie ganzzahlige Company_ID."""
        highest = 0
        for record in self.records:
            value = record.values.get("company_id")
            if value in (None, ""):
                continue
            try:
                highest = max(highest, int(float(value)))
            except (TypeError, ValueError):
                continue
        return highest + 1

    def iter_urls(self) -> Iterator[tuple[str, str]]:
        """(Firmenname, URL) fuer alle Zeilen mit echter http-URL."""
        for record in self.records:
            url = str(record.values.get("url") or "").strip()
            if url.lower().startswith("http"):
                yield str(record.values.get("company") or "").strip(), url

    # -- Schreiben ---------------------------------------------------------

    def write_new_version(
        self,
        new_records: list[Record],
        out_dir: str | Path,
        *,
        version: str | None = None,
        updated: dict[int, Record] | None = None,
    ) -> Path:
        """
        Schreibt eine neue Version: Original kopieren, dann Zeilen ergaenzen
        bzw. bestehende Zeilen ueberschreiben.

        Args:
            new_records: anzuhaengende Zeilen (Company- und Produktzeilen).
            updated:     {0-basierter Datenzeilen-Index: Record} fuer
                         Anreicherungen an bestehenden Zeilen.
            version:     Zielversion, z.B. "2.3". Default: Patch-Bump.

        Returns:
            Pfad der geschriebenen Datei.
        """
        out_dir = Path(out_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        target = out_dir / _bump_filename(self.path.name, version)

        # Kopieren erhaelt Formatierung, Spaltenbreiten und das Legend-Sheet.
        shutil.copyfile(self.path, target)

        workbook = openpyxl.load_workbook(target)
        sheet = workbook[SHEET_MASTER]

        if updated:
            for data_index, record in updated.items():
                excel_row = data_index + 2  # +1 Header, +1 1-basiert
                # Bei Updates werden ALLE Felder geschrieben, auch None.
                # Ein Record aus `updated` ist eine vollstaendige Kopie der
                # Zeile - wuerde None uebersprungen, liessen sich Felder nie
                # LEEREN. Genau das braucht die Reparatur, um einen
                # Seitentitel aus dem URL-Feld zu entfernen.
                # `sheet.cell(..., value=None)` ist in openpyxl ein No-op -
                # die Zuweisung erfolgt nur, wenn value is not None. Zum
                # Leeren muss .value direkt gesetzt werden.
                for fieldname, pos in self.column_map.items():
                    value = record.values.get(fieldname)
                    sheet.cell(row=excel_row, column=pos + 1).value = (
                        None if value == "" else value
                    )

        # NICHT sheet.max_row verwenden: openpyxl zaehlt auch leere, aber
        # formatierte Zeilen mit. v2.2 meldet max_row=983 bei nur 918
        # Datenzeilen - die neuen Zeilen wuerden hinter einer Luecke von
        # ~65 Leerzeilen landen. Stattdessen die letzte Zeile mit Inhalt in
        # der Company-Spalte suchen.
        append_row = _last_data_row(sheet, self.column_map["company"] + 1) + 1
        for record in new_records:
            for fieldname, pos in self.column_map.items():
                value = record.values.get(fieldname)
                if value not in (None, ""):
                    sheet.cell(row=append_row, column=pos + 1, value=value)
            append_row += 1

        workbook.save(target)
        workbook.close()
        log.info("Neue Version geschrieben: %s (+%d Zeilen, %d aktualisiert)",
                 target.name, len(new_records), len(updated or {}))
        return target


def _last_data_row(sheet: Any, column: int) -> int:
    """
    Letzte Zeile mit echtem Inhalt in `column` (1-basiert).

    Laeuft von unten nach oben, weil die Leerzeilen am Ende stehen.
    """
    for row in range(sheet.max_row, 1, -1):
        if str(sheet.cell(row=row, column=column).value or "").strip():
            return row
    return 1


def _block_key(record: Record) -> str:
    """Firmen ohne Company_ID werden ueber den Namen gruppiert."""
    company_id = record.values.get("company_id")
    if company_id not in (None, ""):
        try:
            return f"id:{int(float(company_id))}"
        except (TypeError, ValueError):
            return f"id:{company_id}"
    return f"name:{str(record.values.get('company') or '').strip().lower()}"


def _bump_filename(name: str, version: str | None) -> str:
    """Competitive_Intel_Master_DB_v2.2.xlsx -> ..._v2.3.xlsx"""
    match = VERSION_RE.search(name)
    if version:
        if match:
            return name[: match.start()] + f"_v{version}.xlsx"
        return name.replace(".xlsx", f"_v{version}.xlsx")
    if match:
        major, minor = int(match.group(1)), int(match.group(2))
        return name[: match.start()] + f"_v{major}.{minor + 1}.xlsx"
    stamp = dt.date.today().strftime("%y%m%d")
    return name.replace(".xlsx", f"_enriched_{stamp}.xlsx")


def today_stamp() -> str:
    """Format wie in der Last_Update-Spalte von v2.2."""
    return dt.date.today().strftime("%Y-%m-%d")
