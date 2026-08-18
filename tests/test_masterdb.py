"""Master-DB laden, Zeilenmodell rekonstruieren, versioniert schreiben."""

from __future__ import annotations

import openpyxl
import pytest

from cintel.masterdb import MasterDB, _bump_filename, _last_data_row
from cintel.schema import ROW_KIND_COMPANY, ROW_KIND_PRODUCT, Record


def test_loads_and_skips_trailing_empty_rows(mini_master):
    db = MasterDB(mini_master)
    assert len(db) == 4, "5 Leerzeilen am Ende duerfen nicht mitgezaehlt werden"
    assert len(db.column_map) == 31


def test_blocks_reconstruct_company_product_model(mini_master):
    db = MasterDB(mini_master)
    blocks = {b.company: b for b in db.blocks()}
    assert set(blocks) == {"Makersite", "Autodesk", "Physna"}

    makersite = blocks["Makersite"]
    assert makersite.company_row is not None
    assert makersite.company_row.values["row_kind"] == ROW_KIND_COMPANY
    assert len(makersite.product_rows) == 1
    assert makersite.has_products

    autodesk = blocks["Autodesk"]
    assert autodesk.company_row is not None
    assert not autodesk.has_products, "Autodesk hat keine Produktzeile"

    physna = blocks["Physna"]
    assert physna.company_row is None, "nur Produktzeile vorhanden"
    assert physna.has_products


def test_next_company_id_handles_float_ids(mini_master):
    db = MasterDB(mini_master)
    assert db.next_company_id() == 4


def test_urls_only_returns_real_urls(mini_master):
    db = MasterDB(mini_master)
    blocks = {b.company: b for b in db.blocks()}
    # Autodesk hat nur einen Seitentitel im URL-Feld.
    assert blocks["Autodesk"].urls == []
    assert blocks["Makersite"].urls, "verschmutzte, aber gueltige URL zaehlt"


def test_missing_fields_is_level_aware(mini_master):
    """
    Product_name ist auf der Company-Zeile per Design leer. Eine naive
    Pruefung wuerde jede Firma als luecken behaftet melden.
    """
    db = MasterDB(mini_master)
    blocks = {b.company: b for b in db.blocks()}

    # Makersite HAT ein Produkt -> product_name gilt nicht als Luecke.
    assert "product_name" not in blocks["Makersite"].missing_fields(["product_name"])
    # Autodesk hat keins -> Luecke.
    assert "product_name" in blocks["Autodesk"].missing_fields(["product_name"])
    # Company-Level-Feld wird auf der Company-Zeile geprueft.
    assert "founding_year" not in blocks["Makersite"].missing_fields(["founding_year"])
    assert "founding_year" in blocks["Autodesk"].missing_fields(["founding_year"])


@pytest.mark.parametrize("name,version,expected", [
    ("Competitive_Intel_Master_DB_v2.2.xlsx", None, "Competitive_Intel_Master_DB_v2.3.xlsx"),
    ("Competitive_Intel_Master_DB_v2.2.xlsx", "3.0", "Competitive_Intel_Master_DB_v3.0.xlsx"),
    ("Competitive_Intel_Master_DB_v2.9.xlsx", None, "Competitive_Intel_Master_DB_v2.10.xlsx"),
])
def test_bump_filename(name, version, expected):
    assert _bump_filename(name, version) == expected


def test_write_new_version_appends_without_gap(mini_master, tmp_path):
    """
    Regression: openpyxl meldet auch leere, aber formatierte Zeilen in
    max_row. Ohne _last_data_row landen neue Zeilen hinter einer Luecke.
    """
    db = MasterDB(mini_master)
    new = [
        Record(values={
            "company_id": 4, "company": "Neu AG", "row_kind": ROW_KIND_COMPANY,
            "url": "https://neu.example", "competitor_tier": "Tier 2 – Nachbar",
        }),
        Record(values={
            "company_id": 4, "company": "Neu AG", "product_name": "Neu One",
            "row_kind": ROW_KIND_PRODUCT, "url": "https://neu.example/one",
        }),
    ]
    target = db.write_new_version(new, tmp_path / "out")
    assert target.name == "Competitive_Intel_Master_DB_v2.3.xlsx"

    written = MasterDB(target)
    assert len(written) == 6, "4 vorhandene + 2 neue, keine Leerzeilen dazwischen"
    assert written.records[-1].values["product_name"] == "Neu One"
    assert written.records[-2].values["company"] == "Neu AG"

    # Legend-Sheet muss erhalten bleiben.
    workbook = openpyxl.load_workbook(target)
    assert "Legend_Categories" in workbook.sheetnames


def test_write_new_version_does_not_touch_source(mini_master, tmp_path):
    before = mini_master.read_bytes()
    db = MasterDB(mini_master)
    db.write_new_version(
        [Record(values={"company": "X", "row_kind": ROW_KIND_COMPANY})],
        tmp_path / "out",
    )
    assert mini_master.read_bytes() == before, "Eingangsdatei muss unveraendert sein"


def test_updated_rows_are_written(mini_master, tmp_path):
    db = MasterDB(mini_master)
    original = db.records[0]
    patched = Record(values={**original.values, "location": "Berlin, Germany"})
    target = db.write_new_version([], tmp_path / "out", updated={0: patched})
    written = MasterDB(target)
    assert written.records[0].values["location"] == "Berlin, Germany"


def test_last_data_row_ignores_styled_empty_rows(mini_master):
    workbook = openpyxl.load_workbook(mini_master)
    sheet = workbook["Competitors_All-Master"]
    assert _last_data_row(sheet, 3) == 5, "4 Datenzeilen + Headerzeile"


@pytest.mark.parametrize("name,version,expected", [
    # Explizite Version ersetzt ein vorhandenes Suffix, auch ein untypisches.
    ("Competitive_Intel_Master_DB_v2.2r.xlsx", "2.3",
     "Competitive_Intel_Master_DB_v2.3.xlsx"),
    ("Competitive_Intel_Master_DB_v1.7_short.xlsx", "2.0",
     "Competitive_Intel_Master_DB_v2.0.xlsx"),
    # Ohne Versionssuffix wird angehaengt.
    ("Master.xlsx", "2.3", "Master_v2.3.xlsx"),
])
def test_bump_filename_replaces_odd_version_suffix(name, version, expected):
    """Regression: aus _v2.2r + 2.3 wurde faelschlich _v2.2r_v2.3."""
    assert _bump_filename(name, version) == expected
