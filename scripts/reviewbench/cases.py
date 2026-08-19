"""
Prüffälle mit bekannter Wahrheit.

Jeder Fall ist ein Fehler, der in diesem Projekt wirklich aufgetreten ist.
Gemeinsames Merkmal: der Code laeuft durch, wirft keine Ausnahme, liefert
Exitcode 0 - und ein falsches Ergebnis. Genau die Klasse, die kein Test
findet.

Wichtig: Der Code wird VOLLSTAENDIG uebergeben, nicht als Auszug. Im ersten
Versuch hatte ein gekuerzter Auszug zwei Scheinbefunde erzeugt ("Parameter
ungenutzt", "returncode ungeprueft"), die auf den echten Code nicht zutrafen.
"""

from __future__ import annotations

from dataclasses import dataclass, field


@dataclass
class Fall:
    kennung: str
    titel: str
    klasse: str
    vertrag: str
    code: str
    symptom: str
    ursache: str          # Wahrheit - wird dem Pruefer NICHT gezeigt
    treffer_marker: list[str] = field(default_factory=list)
    # Mindestens ein Marker aus jeder Gruppe muss vorkommen, damit der
    # Befund als Treffer zaehlt. Verhindert Zufallstreffer durch ein
    # einzelnes Allerweltswort.
    treffer_gruppen: list[list[str]] = field(default_factory=list)


FALL_01 = Fall(
    kennung="01_websearch",
    titel="WebSearch wird stillschweigend verweigert",
    klasse="externes Werkzeug, Berechtigung",
    vertrag=(
        "call_claude() ruft die Claude-Code-CLI im Headless-Modus auf. Wenn "
        "tools='WebSearch,WebFetch' uebergeben wird, MUSS das Modell "
        "tatsaechlich im Web suchen koennen. discover_new() nutzt das, um "
        "reale Unternehmen in einem Marktsegment zu finden."
    ),
    code='''
import json, subprocess, shutil

class LLMError(RuntimeError):
    pass

def call_claude(prompt, schema, *, model="sonnet",
                tools="WebSearch,WebFetch", system_prompt="Du bist praezise.",
                timeout=600, max_retries=2):
    """Ruft `claude -p` mit erzwungenem JSON-Schema auf."""
    if shutil.which("claude") is None:
        raise LLMError("claude-CLI nicht gefunden")

    schema_json = json.dumps(schema, ensure_ascii=False)
    last_error = None

    for attempt in range(1, max_retries + 2):
        cmd = [
            shutil.which("claude"), "-p",
            "--output-format", "json",
            "--json-schema", schema_json,
            "--system-prompt", system_prompt,
            "--model", model,
            "--permission-mode", "dontAsk",
            "--no-session-persistence",
        ]
        cmd += ["--tools", tools]

        try:
            proc = subprocess.run(
                cmd, input=prompt, capture_output=True, text=True,
                encoding="utf-8", errors="replace", timeout=timeout,
            )
        except subprocess.TimeoutExpired:
            last_error = LLMError(f"Timeout nach {timeout}s")
            continue

        if proc.returncode != 0 and not proc.stdout.strip():
            last_error = LLMError(f"Exit {proc.returncode}: {proc.stderr[:300]}")
            continue

        try:
            envelope = json.loads(proc.stdout)
        except json.JSONDecodeError:
            last_error = LLMError("kein JSON")
            continue

        if envelope.get("is_error"):
            raise LLMError(str(envelope.get("result"))[:300])

        data = envelope.get("structured_output")
        if data is not None:
            return data
        last_error = LLMError("kein structured_output")

    raise LLMError(f"nach {max_retries + 1} Versuchen erfolglos: {last_error}")


DISCOVERY_SCHEMA = {
    "type": "object", "additionalProperties": False,
    "required": ["companies"],
    "properties": {"companies": {"type": "array", "items": {
        "type": "object", "additionalProperties": False,
        "required": ["name", "homepage"],
        "properties": {"name": {"type": "string"},
                       "homepage": {"type": "string"}}}}},
}

def discover_new(config, known_names, *, model="sonnet", limit=25):
    """Websuche je (Key Category x Sub Category x Region)."""
    out = []
    for key_category in config["key_categories"]:
        for sub_category in config["sub_categories"]:
            for region in config["regions"]:
                if len(out) >= limit:
                    return out
                prompt = (
                    f"Finde reale Unternehmen. Nutze Websuche.\\n"
                    f"Key Category: {key_category}\\n"
                    f"Sub Category: {sub_category}\\n"
                    f"Region: {region}\\n"
                    f"Bereits bekannt (nicht nennen): {', '.join(sorted(known_names)[:400])}"
                )
                result = call_claude(prompt, DISCOVERY_SCHEMA, model=model,
                                     tools="WebSearch,WebFetch")
                gefunden = result.get("companies", []) or []
                for entry in gefunden:
                    out.append({"name": entry["name"],
                                "homepage": entry["homepage"]})
                log.info("Discovery %s/%s/%s: %d Kandidaten",
                         key_category, sub_category, region, len(gefunden))
    return out
''',
    symptom=(
        "Im Betrieb liefert discover_new() ueber ALLE Suchzellen hinweg "
        "konstant 0 Kandidaten. Es wird keine Ausnahme geworfen, der Exitcode "
        "ist 0, und die Antwort ist schema-gueltig."
    ),
    ursache=(
        "--permission-mode dontAsk verweigert Werkzeuge, statt sie "
        "stillschweigend zu erlauben. Ohne zusaetzliches --allowedTools wird "
        "WebSearch abgelehnt; das Modell liefert daraufhin eine leere, aber "
        "schema-gueltige Antwort."
    ),
    treffer_gruppen=[
        ["permission", "berechtigung", "dontask", "don't ask", "erlaub",
         "allowedtools", "allow", "freigab", "verweiger", "denied", "denial"],
        ["websearch", "web search", "websuche", "tool"],
    ],
)


FALL_02 = Fall(
    kennung="02_cell_none",
    titel="Feld laesst sich nicht leeren",
    klasse="Bibliotheks-Semantik, stiller No-op",
    vertrag=(
        "repariere_zeilen() bereinigt eine Excel-Tabelle. Enthaelt das "
        "URL-Feld keinen echten Link, sondern Fliesstext (z.B. einen "
        "Seitentitel), MUSS das Feld GELEERT und der Fremdinhalt in die "
        "Bemerkungen verschoben werden. Die Eingangsdatei bleibt unveraendert; "
        "geschrieben wird eine Kopie."
    ),
    code='''
import re, shutil
import openpyxl

URL_RE = re.compile(r"https?://[^\\s|,;]+", re.IGNORECASE)

def erste_url(wert):
    """Zieht die erste echte http(s)-URL aus einem verschmutzten Feld."""
    if not wert:
        return ""
    treffer = URL_RE.search(str(wert))
    return treffer.group(0).rstrip("/.,;|") if treffer else ""


def repariere_zeilen(records, spalten_map):
    """
    Erzeugt {Zeilenindex: bereinigter Record}.
    spalten_map bildet Feldnamen auf 0-basierte Spaltenpositionen ab.
    """
    updates = {}
    for index, record in enumerate(records):
        neu = dict(record)
        geaendert = False

        roh = str(neu.get("url") or "").strip()
        if roh:
            url = erste_url(roh)
            if not url:
                # Kein Link enthalten -> Feld leeren, Inhalt retten
                notiz = str(neu.get("remarks") or "").strip()
                marker = f"[aus URL-Feld uebernommen] {roh}"
                neu["remarks"] = f"{notiz}\\n{marker}".strip() if notiz else marker
                neu["url"] = None
                geaendert = True
            elif url != roh.rstrip("/"):
                neu["url"] = url
                geaendert = True

        if geaendert:
            updates[index] = neu
    return updates


def schreibe_neue_fassung(quelle, ziel, updates, spalten_map):
    """Kopiert die Datei und traegt die bereinigten Zeilen ein."""
    shutil.copyfile(quelle, ziel)
    wb = openpyxl.load_workbook(ziel)
    ws = wb["Competitors_All-Master"]

    for daten_index, record in updates.items():
        excel_zeile = daten_index + 2          # +1 Kopfzeile, +1 1-basiert
        for feldname, position in spalten_map.items():
            wert = record.get(feldname)
            ws.cell(row=excel_zeile, column=position + 1, value=wert)

    wb.save(ziel)
    wb.close()
    return ziel
''',
    symptom=(
        "Nach dem Lauf meldet die anschliessende Pruefung weiterhin "
        "'URL-Feld enthaelt keine URL' fuer genau dieselben Zeilen. Der "
        "Reparaturbericht listet die Aenderung aber als durchgefuehrt auf, "
        "und es gibt keinen Fehler."
    ),
    ursache=(
        "openpyxl: Worksheet.cell(..., value=None) ist ein No-op - die "
        "Zuweisung erfolgt nur, wenn value is not None. Zum Leeren muss "
        ".value direkt gesetzt werden."
    ),
    treffer_gruppen=[
        ["none", "null", "leer"],
        ["cell(", "cell (", ".value", "openpyxl", "zuweis", "no-op", "noop",
         "ueberschrieben", "überschrieben", "nicht gesetzt", "ignor"],
    ],
)


FALL_03 = Fall(
    kennung="03_robots",
    titel="Offene Seiten gelten als gesperrt",
    klasse="Bibliotheks-Semantik + Netz",
    vertrag=(
        "ist_erlaubt() prueft die robots.txt einer Domain und gibt True "
        "zurueck, wenn der Crawler die URL abrufen darf. Enthaelt robots.txt "
        "'User-agent: *' mit leerem 'Disallow:', ist ALLES erlaubt. Der "
        "Crawler soll robots.txt respektieren, aber keine Seite faelschlich "
        "aussperren."
    ),
    code='''
import urllib.robotparser
from urllib.parse import urljoin, urlparse
import requests

USER_AGENT = "cintel/0.1 (+https://example.org/bot)"

class Crawler:
    def __init__(self, timeout=20, respect_robots=True):
        self.timeout = timeout
        self.respect_robots = respect_robots
        self._robots = {}
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": USER_AGENT})

    def ist_erlaubt(self, url):
        """Darf diese URL abgerufen werden?"""
        if not self.respect_robots:
            return True

        parsed = urlparse(url)
        origin = f"{parsed.scheme}://{parsed.netloc}"

        if origin not in self._robots:
            parser = urllib.robotparser.RobotFileParser()
            parser.set_url(urljoin(origin, "/robots.txt"))
            try:
                parser.read()
            except Exception:
                parser = None      # robots.txt nicht lesbar -> nicht blockieren
            self._robots[origin] = parser

        parser = self._robots[origin]
        if parser is None:
            return True
        try:
            return parser.can_fetch(USER_AGENT, url)
        except Exception:
            return True

    def hole(self, url):
        """Ruft eine Seite ab, sofern robots.txt es erlaubt."""
        if not self.ist_erlaubt(url):
            return {"url": url, "status": 0,
                    "fehler": "durch robots.txt untersagt"}
        antwort = self.session.get(url, timeout=self.timeout,
                                   allow_redirects=True)
        return {"url": url, "status": antwort.status_code,
                "text": antwort.text}
''',
    symptom=(
        "Mehrere Firmenseiten werden mit 'durch robots.txt untersagt' "
        "abgelehnt. Ruft man deren robots.txt im Browser auf, steht dort "
        "'User-agent: *' und 'Disallow:' (leer) - also alles erlaubt."
    ),
    ursache=(
        "RobotFileParser.read() holt robots.txt mit dem urllib-Default-"
        "User-Agent. WAFs wie Cloudflare antworten darauf mit 403, und ein "
        "403 laesst RobotFileParser disallow_all setzen. robots.txt muss "
        "ueber die eigene Session mit korrektem User-Agent geholt und dann "
        "per parse() eingelesen werden."
    ),
    treffer_gruppen=[
        ["read()", "parser.read", "urllib", "default", "eigener user-agent",
         "session", "user-agent", "user agent"],
        ["403", "disallow_all", "disallow all", "gesperrt", "blockiert",
         "abgewiesen", "cloudflare", "waf", "bot"],
    ],
)


ALLE_FAELLE = [FALL_01, FALL_02, FALL_03]
